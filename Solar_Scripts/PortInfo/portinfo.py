from openpyxl import Workbook , load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from orionsdk import SwisClient
import requests , abc

class ExcelSheet ( ):

	'''
	Class name    : ExcelSheet
	Class Purpose : To handle reading and writing to Excel spreadsheet
	'''

	def __init__                ( self , workbook ):

		'''
		Method name: __init__
		Method Purpose: To initialize an excel instance

		Parameters:
			- workbook (string): The name of the workbook

		Returns: None
		'''
		
		self.workbookName = workbook
		try:
			self.workbook = self.openWorkbook ( workbook )
		except Exception as detail:
			self.workbook = Workbook ( workbook )
			
		self.worksheet = self.getWorkbookActiveSheet ( )
		self.columns   = self.worksheet.max_column
		self.rows      = self.worksheet.max_row

	def readFullWorkbook        ( self , numRows , numColumns , sheet , startRow=1 ):

		'''
		Method name: readFromWorkbook
		Method Purpose: To read from the inputted workbook

		Parameters:
			- numRows (integer): The number of rows to read
			- numColumns (integer): The number of columns to read
			- startRow (integer): The row to start reading from
			- sheet (unknown): The worksheet to read from

		Returns: A list of all data
		'''

		spreadsheet_data = []
		for sheet_row in range ( *self.__getReadRange ( sheet.max_row , numRows , startRow ) ):
			rowData = self.readRowFromWorkbook ( sheet_row , numColumns , sheet )
			spreadsheet_data.append ( rowData )
		return spreadsheet_data

	def readRowFromWorkbook     ( self , rowNum , numColumns , sheet ):

		'''
		Method name: readRowFromWorkbook
		Method Purpose: To read an individual row from Workbook

		Parameters:
			- rowNum (integer): The numbered row to read
			- numColumns (integer): The number of columns to read from row
			- sheet (unknown): The sheet that we are reading from

		Returns: A list of the data in the row
		'''

		rowData = 	[
						sheet.cell  (
							row=rowNum if rowNum < sheet.max_row + 1 else sheet.max_row, \
							column=column
						).value for column in range ( \
							*self.__getReadRange ( sheet.max_column , numColumns )
						)
					]
		return rowData

	def writeToWorkbook         ( self , sheet , row , rowData , columnStart=0  ):

		'''
		Method name: writeToWorkbook
		Method Purpose: To write to the inputted workbook

		Parameters: 
			- sheet (WorkSheet): The worksheet we are writing to
			- row (integer): The row number to write to
			- rowData (list): A list of data for row
			- columnStart (integer): The starting spot to place the column

		Returns: None
		'''

		colNum = columnStart
		for counter , data in enumerate ( rowData ):
			sheet.cell ( row=row , column=colNum ).value = data
			sheet.cell ( row=row , column=colNum ).alignment = Alignment ( wrap_text = False )
			colNum = colNum + 1	
		self.saveWorkbook ( )
		
	def saveWorkbook            ( self ):
	
		'''
		Method name: saveWorkBook
		Method Purpose: To save the workbook

		Parameters: None

		Returns: None
		'''
	
		self.workbook.save ( self.workbookName )
		
	def removeSheetFromWorkbook ( self , sheetName ):

		'''
		Method name: removeSheetFromWorkbook
		Method Purpose: To remove an inputted sheet from the inputted workbook

		Parameters:
			- sheetName (string): The name of the sheet to remove

		Returns: None
		'''

		try:
			self.workbook.remove ( self.workbook.get_sheet_by_name ( sheetName ) )
		except:
			print ( "The sheet is unable to be deleted because the name is incorrect." )

	def __getReadRange          ( self , maxCell , inputCell , startCell=1 ):

		'''
		Method name: __findReadRange
		Method Purpose: To find the range being read from Excel Worksheet;
						either the row or column range

		Parameters:
			- maxCell (integer): The max number of cells that can be read
			- startCell (integer): The starting cell (row or column)
			- inputCell (integer): The user inputted range to be read

		Returns: A tuple containing the read range
		'''

		cellRange = (
				startCell   if startCell < inputCell and startCell > 0
							else 1 ,
				inputCell + 1 if inputCell  < maxCell + 1 and inputCell < maxCell + 1 \
							else maxCell + 1
			)
		return cellRange

	def getWorkbookActiveSheet  ( self ):

		'''
		Method name: __getWorkbookActiveSheet
		Method Purpose: Get the Workbooks active sheet

		Parameters: None

		Returns: The active sheet of the workbook
		'''

		return self.workbook.active

	def openWorkbook          ( self , workbookName ):

		'''
		Method name: _openWorkbook
		Method Purpose: To open the workbook if it is not already

		Parameters:
			- workbookName (string): The name of the workbook to open

		Returns: None
		'''

		try:
			workbook = load_workbook ( workbookName )
		except FileNotFoundError:
			raise Exception ( "The file name entered could not be found in the file system." )
		else:
			return workbook
			
	def setRowHeight ( self , rowNum , height , worksheet ):
	
		'''
		Method name: setRowHeight
		Method Purpose: To set the row height

		Parameters:
			- rowNum (integer):The row we are adjusting
			- height (integer): The new height of the rwo
			- worksheet (worksheet): The worksheet we are adjusting

		Returns: None
		'''
		
		worksheet.row_dimensions [ rowNum ].height = height
	
	def setColumnWidth ( self , colNum , width , worksheet ):
	
		'''
		Method name: setColumnWidth
		Method Purpose: To set the column width

		Parameters:
			- colNum (integer): The column we are adjusting
			- width (integer): The new width of the column
			- worksheet (worksheet): The worksheet we are adjusting

		Returns: None
		'''
		
		worksheet.column_dimensions [ get_column_letter ( colNum ) ].width = width
		
	def getRowHeight ( self , rowNum , worksheet ):
	
		'''
		Method name: setRowHeight
		Method Purpose: To set the row height

		Parameters:
			- rowNum (integer):The row we are adjusting
			- worksheet (worksheet): The worksheet we are adjusting

		Returns: None
		'''
		
		height = worksheet.row_dimensions [ rowNum ].height
		return height
		
	def getColumnWidth ( self , colNum , worksheet ):
	
		'''
		Method name: setColumnWidth
		Method Purpose: To set the column width

		Parameters:
			- colNum (integer): The column we are adjusting
			- worksheet (worksheet): The worksheet we are adjusting

		Returns: None
		'''

		width = worksheet.column_dimensions [ get_column_letter ( colNum ) ].width
		return width

class SolarwindsEntity ( abc.ABC ):

	'''
	Class name: SolarwindsEntity
	Class Purpose: To serve as base class for all Solarwinds entities
	'''
		
	def __init__ ( self , domain , username , password ):
	
		'''
		Method name: __init__
		Method Purpose: To start a solarwinds instance

		Parameters:
			- domain (string): The domain to log into
			- username (string): The username to log into the system
			- password (string): The password for the associated username

		Returns: None
		'''
		
		# the server is unverified --> allow without warning
		verify = False
		if not verify:
			from requests.packages.urllib3.exceptions import InsecureRequestWarning
			requests.packages.urllib3.disable_warnings ( InsecureRequestWarning )	
		self._solarwinds = SwisClient ( domain , username , password )
		
class PortEntity ( SolarwindsEntity ):

	'''
	Class name: PortEntity
	Class Purpose: To obtain individual port details from Solarwinds
	'''

	def __init__ ( self , domain , username , password , ipAddress , nodeName=None ):

		'''
		Method name: __init__
		Method Purpose: To initialize a port detail instance

		Parameters:
			- ipAddress (string): The ip address the port is bound to
			- username (string): The username of the user in solarwinds
			- password (string): The password of the user in solarwinds

		Returns: None
		'''

		super ( ).__init__ ( domain , username , password )
		self.__ipAddress = ipAddress
		self.__nodeName  = nodeName
	
	def getPortInfo ( self ):

		'''
		Method name: getPortInfo
		Method Purpose: To get the important port information from Solarwinds

		Parameters: None

		Returns: A dictionary of the details of the query if it is found
				 'None' if results are not found
		'''
		
		searchCondition = "e.IPAddresses.IPAddress='{}'".format ( self.__ipAddress ) \
						  if self.__nodeName == None else \
						  "(e.IPAddresses.IPAddress='{}' and e.Ports.Port.Node.Caption='{}')".format ( self.__ipAddress , self.__nodeName )
		portQueryResults = self._solarwinds.query  ( 	
														"""
														SELECT
															e.Ports.Port.Name,
															e.Ports.Port.PortDescription,
															e.Ports.Port.Speed,
															e.Ports.Port.Duplex,
															e.Ports.Port.Node.Caption
														FROM
															Orion.UDT.Endpoint e
														WHERE
															e.Ports.ConnectionType=1 and {}
														""".format ( searchCondition )
													)

		return portQueryResults [ 'results' ]

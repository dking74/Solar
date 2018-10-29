from openpyxl import Workbook
from orionsdk import SwisClient
from abc      import ABC, abstractmethod
import openpyxl, requests, sys , signal , os

class IntelligridMig  ( ):

    '''
        Class name    : IntelligridMig
    
        Class Purpose : To migrate entities from Excel to Solarwinds
    '''

    def __init__         ( self , domain , inputFile ):

        '''
            Method name    : __init__
        
            Method Purpose : To initialize the SolarWinds migration instance
        
            Parameters     :
                -inputFile : The file to be loaded
        
            Returns        : None
        '''

        # the server is unverified --> allow without warnings
        verify = False
        if not verify:
            from requests.packages.urllib3.exceptions import InsecureRequestWarning
            requests.packages.urllib3.disable_warnings ( InsecureRequestWarning )

        # login to solarwinds and load the worksheet
        self.__loginSolar ( domain )
        self._inputFile = inputFile

    def __loginSolar     ( self , domain ):

        '''
            Method name    : loginSolar
        
            Method Purpose : To create an instance of solarwinds
        
            Parameters     : None
        
            Returns        : None
        '''

        valid = False        

        # get the username and password from the user until it is valid
        while not valid:

            username = input ( "Username: " )
            password = input ( "Password: " )
            print            (     "\n"     )

            try: 
                self._solarwinds = SwisClient (      domain , username , password      )
                self._solarwinds.query        ( "SELECT AccountID FROM Orion.Accounts" )
                valid = True

            except requests.exceptions.HTTPError:
                valid = False

    def __loadWorksheet  ( self , file ):
    

        '''
            Method name    : load_workbook
        
            Method Purpose : To load the workbook/worsheet
        
            Parameters     : None
        
            Returns        : None
        '''

        # get the workbook and worksheet locaclly
        self._intelligridBook  = openpyxl.load_workbook ( file )
        self._intelligridSheet = self._intelligridBook['Site ID']

    def _setupBaseGroups ( self , baseGroup ):

        '''
            Method name     : _setupBaseGroups
        
            Method Purpose  : To setup the the base groups for adding
        
            Parameters      :
                - baseGroup : The base-level group to create
        
            Returns         : None
        '''

        # get top level group
        self._baseGroupID, uri = self.createGroup  ( baseGroup , "This is the base-level group" , [] )

    def _removeGroupsFromList(self, removeList, *groups):

        """Remove groups from list"""
        for group in groups:
            try:
                removeList.remove(group)
            except ValueError:
                try:
                    groupName = " " + group
                    removeList.remove(groupName)
                except ValueError:
                    pass
        return removeList
		
    def readWorkbook     ( self , baseGroup ):

        '''
            Method name    : read_workbook
        
            Method Purpose : To read the inputted workbook
        
            Parameters     :
                -baseGroup : The name given to the base-level group
        
            Returns        : None
        '''

        def pauseExecution ( sig , frame ):
    
            '''
                Method name    : pauseExecution
            
                Method Purpose : To pause the execution when reading workbook
            
                Parameters     : 
                    - sig      : The signal received
                    - frame    : The frame received
            
                Returns        : None
            '''

            if sig == signal.SIGINT:
                in_val = input ( "Continue? Y/N:\t" )

                if   in_val == "Y" or in_val == "y": pass
                elif in_val == "N" or in_val == "n": 
                    os.kill ( os.getpid ( ) , signal.SIGABRT )

        # setup a signal handler so that users can pause execution
        signal.signal ( signal.SIGINT , pauseExecution )

        # setup the base level groups and read the file
        self._setupBaseGroups            (    baseGroup    )
        self.__loadWorksheet             ( self._inputFile )
        existingList = self.getGroupList (    baseGroup    )

        # iterate through every row in workbook
        for ROW in range ( 3 , self._intelligridSheet.max_row + 1 ):

            # get the column info from row
            legacy_loc = self._intelligridSheet.cell ( row=ROW , column=1  ).value
            site_id    = self._intelligridSheet.cell ( row=ROW , column=5  ).value
            loc_name   = self._intelligridSheet.cell ( row=ROW , column=7  ).value
            division   = self._intelligridSheet.cell ( row=ROW , column=8  ).value
            owning_co  = self._intelligridSheet.cell ( row=ROW , column=9  ).value
            asset_type = self._intelligridSheet.cell ( row=ROW , column=10 ).value
            latitude   = self._intelligridSheet.cell ( row=ROW , column=11 ).value
            longitude  = self._intelligridSheet.cell ( row=ROW , column=12 ).value
            address    = self._intelligridSheet.cell ( row=ROW , column=13 ).value
            loc_id     = self._intelligridSheet.cell ( row=ROW , column=14 ).value
            emprv_dist = self._intelligridSheet.cell ( row=ROW , column=15 ).value
            prim_dist  = self._intelligridSheet.cell ( row=ROW , column=16 ).value 

            # determine if the nodes exist
            legacy_info = self.detExistingNode ( legacy_loc )
            site_info   = self.detExistingNode (   site_id  )

            # if one of the nodes exist
            if legacy_info or site_info:

                # determine if the nodes have a group in the base group
                group_id                      = 0
                legG_exists = sitG_exists     = False
                lgroup      = sgroup          = ""
                lID         = sID             = None
                legG_exists, lgroup, lID      = self.getNodeContain ( legacy_loc , existingList ) 
                if site_id != legacy_loc:
                    sitG_exists, sgroup, sID  = self.getNodeContain ( site_id    , existingList )

                # create a filter for what to include in the group
                dynamicQuery = self.createFilter   ( 
                                                        "Orion.Nodes"                                                           ,
                                                        self.__formNameString ( legacy_info , site_info , legacy_loc , site_id ),
                                                        "StartsWith"                                                            ,
                                                        [ 
                                                            { 'Caption': site_id    },
                                                            { 'Caption': legacy_loc }
                                                        ] 
                                                    )

                if lgroup[0] == ' ':
                    lgroup
                if   ( legG_exists and sitG_exists ) and ( lID == sID ):
                    existingList = self._removeGroupsFromList(existingList, lgroup)
                    self.updateGroup    ( lID , loc_name , loc_name )
                    group_id = lID

                elif ( legG_exists and sitG_exists ) and ( lID != sID ):
                    existingList = self._removeGroupsFromList(existingList, lgroup, sgroup)
                    self.deleteGroup      (          sgroup           )
                    self.updateGroup      ( lID , loc_name , loc_name )
                    self.deleteDefinition (           lID             )
                    self.updateDefinition (      lID , dynamicQuery   )
                    group_id = lID

                elif ( legG_exists and not sitG_exists ):
                    existingList = self._removeGroupsFromList(existingList, lgroup)
                    self.updateGroup      ( lID , loc_name , loc_name )
                    self.deleteDefinition (            lID            )
                    self.updateDefinition (      lID , dynamicQuery   )
                    group_id = lID

                elif ( sitG_exists and not legG_exists ):
                    existingList = self._removeGroupsFromList(existingList, sgroup)
                    self.updateGroup      ( sID , loc_name , loc_name )
                    self.deleteDefinition (            sID            )
                    self.updateDefinition (      sID , dynamicQuery   )
                    group_id = sID

                else:
                    group_id, group_uri = self.createGroup      ( loc_name , loc_name , dynamicQuery )
                    if group_id != None:  self.createDefinition ( 
                                                                    self._baseGroupID , \
                                                                    [
                                                                        {
                                                                            'Name'      : loc_name,
                                                                            'Definition': group_uri
                                                                        }
                                                                    ]
                                                                )

                # update the properties
                if group_id != None:
                    properties = self.createGroupProp ( division, owning_co, asset_type, latitude, longitude, \
                                                         address, loc_id, emprv_dist, prim_dist               )
                    self.updateGroupProps             (              group_id , **properties                  )
                    self.updateMapPoint               (           group_id , latitude , longitude             )
            
            # if no node exists by the name
            else:
                print ( "There were no nodes matching: {} or {}".format ( legacy_loc , site_id ) )
                    
        for group in existingList:
            self.deleteGroup ( group )

    # def createSheetHeaderList ( self , worksheet ):

    #   for col in range ( worksheet.max_row ) + 1:
    #       self._headerList.append ( worksheet.cell ( row=1 , column=col ).value )

    # def getRowData     ( self , worksheet , ROW ):

        # for col in range ( worksheet.max_row ) + 1:
        #   data.append ( worksheet.cell ( row=ROW , column=col ).value )
        #

        # 
    
    def detExistingNode   ( self , check_str ):

        '''
            Method name     : detExistingNode
        
            Method Purpose  : To determine if a node is present
        
            Parameters      : 
                - check_str : The string of the node to look for
        
            Returns         :
                - True      : If there are nodes found
                - False     : If there are no nodes found
        '''

        query_res = { 'results': [] }
        if check_str != '':
            query_res = self._solarwinds.query (    """
                                                    SELECT
                                                        Caption,
                                                        Uri
                                                    FROM 
                                                        Orion.Nodes 
                                                    WHERE
                                                        Caption
                                                    LIKE '{}%'
                                                    """.format ( str ( check_str ).lower ( ) ) 
                                                )  

        # if there are nodes, return True
        if len ( query_res [ 'results' ] ) > 0: return True
        else                                  : return False         

    def __formNameString  ( self, legacyInfo , siteInfo , legacyStr , siteStr ):

        '''
            Method name    : formNameString
        
            Method Purpose : To build a string for the dynamic query name
        
            Parameters     :
        
            Returns        :
        '''

        names = []
        if legacyInfo: names.append ( legacyStr )
        if siteInfo  : names.append (  siteStr  )
        nameString = ", ".join ( names )
        return nameString 

    def updateGroupProps  ( self , entity_id , **properties ):

        '''
            Method name      : updateCustProps
        
            Method Purpose   : To add properties to a group
        
            Parameters       :
                - entity_id  : The entity to update
                - properties : The properties to give entity
        
            Returns          : None
        '''

        # get the Uri of the entity
        result_uri = self._solarwinds.query (   """
                                                SELECT
                                                    Uri
                                                FROM
                                                    Orion.Container
                                                WHERE
                                                    ContainerID='{}'
                                                """.format ( entity_id )
                                            )

        # pull the uri directly
        uri = result_uri [ 'results' ][ 0 ][ 'Uri' ]

        # update the entity with inputted properties
        try:
            self._solarwinds.update ( str(uri)+ '/CustomProperties' , **properties )

        except Exception as detail:
            print ( str(detail)  + " \nUnable to update custom properties for id: {}".format ( entity_id ) )

    def updateNodeProp    ( self , node_name , cust_prop , value ):
        
        '''
            Method name      : updateNodeCustProp
        
            Method Purpose   : To add property to a node
        
            Parameters       :
                - node_name  : The node to update
                - cust_prop  : The custom property to update
                - value      : Value to give custom property
        
            Returns          : None
        '''

        # make sure value is not empty
        if value != "":

            # get the Uri of the entity
            result_uri = self._solarwinds.query (   """
                                                    SELECT
                                                        Uri
                                                    FROM
                                                        Orion.Nodes
                                                    WHERE
                                                        Caption
                                                    LIKE 
                                                        '{}%'
                                                    """.format ( node_name )
                                                )

             # create dictionary for property
            properties = { cust_prop : value }

            # pull the uri directly
            for result in result_uri [ 'results']:  
                uri = result [ 'Uri' ]

                # update the entity with inputted properties
                try:
                    self._solarwinds.update ( uri + '/CustomProperties' , **properties )
                    print                   ( "Updated Custom Property for node: %s" % node_name )
                except Exception:
                    print ( "Unable to update Custom Property for node: %s" % node_name )

    def updateDefinition  ( self , id , *definition ):

        '''
            Method name      : updateDefinition
        
            Method Purpose   : To update the definition of a group
        
            Parameters       :
                - id         : The id of the group to update
                - definition : The new definition for the group
        
            Returns          : None
        '''

        # get the definition id for the appropriate container
        currentDef = self._solarwinds.query (
                    """
                    SELECT
                        DefinitionID
                    FROM
                        Orion.ContainerMemberDefinition
                    WHERE
                        ContainerID='{}'
                    """.format ( id )
                                            
                )

        # see if there are any results and then update the definition
        if len ( currentDef [ 'results' ] ) > 0:
            newDef = self._solarwinds.invoke (
                        'Orion.Container',
                        'UpdateDefinition',
                        currentDef [ 'results' ][ 0 ][ 'DefinitionID' ],
                        definition [ 0 ]

                    )
        else:
            self.createDefinition ( id , definition [ 0 ] )

    def updateGroup       ( self , id , newName , newDescription ):

        '''
            Method name          : updateGroup
        
            Method Purpose       : To update the group with new properties
        
            Parameters           :
                - id             : The id of the group to update
                - newName        : The new name to give the group
                - newDescription : The new description to give the group
        
            Returns              : The name of the new group
        '''

        try: 
            update = self._solarwinds.invoke (
                        "Orion.Container",
                        "UpdateContainer",
                        id,
                        newName,
                        "Core",
                        60,
                        0,
                        newDescription,
                        True
                    )
        except Exception:
            print ( "Unable to update group due to internal error")
            return None
        return newName

    def updateMapPoint    ( self , group_id , latitude , longitude ):

        '''
            Method name     : updateMapPoint
        
            Method Purpose  : To update a current map point for a group
        
            Parameters      :
                - group_id  : The id of the group to modify
                - latitude  : The latitude coordinate
                - longitude : The longitude coordinate
        
            Returns         : None
        '''

        # see if there are any map points for the group
        mapPointView = self._solarwinds.query   ( 
                            """
                            SELECT
                                Uri
                            FROM
                                Orion.WorldMap.Point
                            WHERE
                                PointId='{}'
                            """.format ( group_id )
                        )
        props = {
            'Latitude'  : latitude,
            'Longitude' : longitude
        }

        # update the map point, otherwise create a new one
        if latitude != None:
            try:
                uri = mapPointView [ 'results' ][ 0 ][ 'Uri' ]
                self._solarwinds.update ( uri , **props )
            except Exception:
                self.createMapPoint ( group_id , latitude , longitude )

    def createFilter      ( self , filterType , nameInput , verbSearch , *entityList ):

        '''
            Method name      : createFilter
        
            Method Purpose   : To create a filter for members in group
        
            Parameters       :
                - filterType : What entity we are filtering on
                - nameInput  : The name to give the dynamic query
                - verbSearch : The verb that will be used to query entities 
                    *example - StartsWith or Contains (sql parse instructions)
                - entityList : The list of dictionary definitions to query on 
        
            Returns          : The filter to apply
        '''

        # create a string that is used for parsing entities
        filterList   = []
        for dictionary in entityList [ 0 ]:
            key        = list ( dictionary.keys   ( ) ) [ 0 ]
            value      = list ( dictionary.values ( ) ) [ 0 ]
            if value != "": 
                tempString = "{}({},'{}')".format ( verbSearch , key , value ) 
                filterList.append ( tempString )
        if filterList != []: filterString = ' or '.join ( filterList )

        # create the actual filter
        name        = nameInput
        filtered    = "filter:/{}[{}]".format ( filterType , filterString )
        applyFilter = [ { 'Name': name , 'Definition' : filtered } ]
        return applyFilter

    def createGroupProp   ( self , division=None, owning_co=None, asset_type=None, latitude=None, longitude=None,
                             address=None, loc_id=None , emprv_dist=None , prim_dist=None ):

        '''
            Method name      : defCustProp
        
            Method Purpose   : To create a list of custom properties for Solarwinds
        
            Parameters       :
                - division   : The division of the area
                - owning_co  : The company that owns the resource
                - asset_type : What kind of asset it is 
                - latitude   : The latitude location of resource
                - longitude  : The longitude location of resource
                - address    : The address (or notes) for area
                - loc_id     : The id given to location
                - emprv_dist : EMPRV info
                - prim_dist  : PRIMAVERA info
        
            Returns          : The dictionary containing the properties
        '''

        # build dictionary based on info
        cust_prop = {

            'Division'        : division,
            'Owning_Company'  : owning_co,
            'Asset_Type'      : asset_type,
            'Latitude'        : latitude,
            'Longitude'       : longitude,
            'Address'         : address,
            'Location_ID'     : loc_id,
            'EMPRV_Dist'      : emprv_dist,
            'Primavera_Dist'  : prim_dist
        }

        # iterate through each custom property and determine if there is an empty value
        # if so --> keep a list of every item to be deleted
        del_items = []
        for k, v in cust_prop.items ( ):
            if v == '' or v == None:
                del_items.append ( k )

        # delete all unnecessary items
        for item in del_items:
            cust_prop.pop ( item )
        return cust_prop

    def createMapPoint    ( self, group_id , latitude , longitude ):

        '''
            Method name      : updateMapPoint
        
            Method Purpose   : To update the map location of a group
        
            Parameters       :
                - group_id   : The id of the group
                - latitude   : The latitude location
                - longitude  : The longitude location
        
            Returns          : 
                - True       : If successful
                - Fale       : If failing
        '''

        # create the properties
        properties = {
             'Instance'  : 'Orion.Groups',
             'InstanceID': group_id,
             'Latitude'  : latitude,
             'Longitude' : longitude
        }

        # create the map point based on properties
        try:
            info = self._solarwinds.create ( 'Orion.WorldMap.Point', **properties )
        except Exception as detail:
            print ( str(detail) + "\nUnable to create map point for id: {}".format ( group_id ) )

    def createCustProps   ( self , entity_type , **properties ):

        '''
            Method name       : createCustProps
        
            Method Purpose    : To create a custom property in Solarwinds
                              : for each item in the list
        
            Parameters        :
                - entity_type : The type that is being given the prop.
                - properties  : The list of properties to add
        
            Returns           : None
        '''

        # iterate through each property
        for prop, descr in properties.items ( ):

            # check to make sure property is not there already
            prop_find = self._solarwinds.query (    """
                                                    SELECT
                                                        N.Description
                                                    FROM
                                                        Orion.CustomProperty N
                                                    WHERE
                                                        N.Description='{}'
                                                    """.format ( descr )
                                                )

            # if there isn't, add the custom properties to solarwinds
            if len ( prop_find [ 'results' ] ) == 0:

                # determine type and size of property
                if prop == 'Latitude' or prop == 'Longitude':
                    type_prop = 'double'
                    size      = None
                else:
                    type_prop = 'string'
                    size      = 100

                try:
                    self._solarwinds.invoke (   
                                                entity_type,
                                                'CreateCustomProperty',
                                                prop,
                                                descr,
                                                type_prop,
                                                size,
                                                None,
                                                None,
                                                None,
                                                None,
                                                None,
                                                None
                                            )
                except Exception:
                    print ( "Unable to create custom property." )
            else:
                print ( "Custom property already exists for: {}".format ( prop ) )

    def createDefinition  ( self , id_num , *args ):

        '''
            Method name    : addDefinitions
        
            Method Purpose : To add entities to a group
        
            Parameters     :
                - id       : The container id
                - args     : The entities to add
        
            Returns        : None
        '''

        if len ( *args ) == 1:
            definition = 'AddDefinition'
            argument   = args [ 0 ]
        else:
            definition = 'AddDefinitions'
            argument   = args

        # update the container
        try:
            self._solarwinds.invoke ( 
                "Orion.Container",
                definition,
                id_num,
                *argument
            )

        except Exception:
            print ( "Unable to add to container: {}".format ( id_num ) )

    def createGroup       ( self , group_name, description, *nodes ):

        '''
            Method name        : createGroup
        
            Method Purpose     : To create a group in SolarWinds
        
            Parameters         :
                - group_name   : The name given to the group
                - description  : The description of the group
                - *nodes       : The nodes to add
        
            Returns            :
                - group_info   : The group created in a list
                - id_num       : The id of the group
        '''

        # get container info to see if a group is present
        container = self.getGroupInfo ( group_name )

        # there are no results, add the group
        if container == "None":
            try:
                id_num = self._solarwinds.invoke (
                            "Orion.Container",
                            "CreateContainer",
                            group_name,
                            "Core",
                            60,
                            0,
                            description,
                            True,
                            *nodes
                        )

            # check if the group is unable to be added
            except ( requests.exceptions.HTTPError, Exception ):
                print ( "Unable to create group because invalid members were detected." )
                return None, None

            # otherwise, add the group to the list and return
            else:
                print ( "Group {} created!".format ( group_name.upper ( ) ) )
                results = self.getGroupInfo ( group_name )
                return results [ 'results' ][ 0 ][ 'ContainerID' ], \
                       results [ 'results' ][ 0 ][     'Uri'     ] 
        else:
            print ( "Group {} already exists".format ( group_name.upper ( ) ) )
            return container [ 'results' ][ 0 ][ 'ContainerID' ], \
                   container [ 'results' ][ 0 ][      'Uri'    ]

    def deleteGroup       ( self , name ):

        '''
            Method name    : deleteEntity
        
            Method Purpose : To delete an entity from Solarwinds
        
            Parameters     :
                - name     : The name of the entity
        
            Returns        :
                - True     : If the entity is deleted
                - False    : If unable to locate entity
        '''

        # get the container ID
        container = self.getGroupInfo ( name )

        # error check to make sure we can delete
        if container == "None":
            print ( "No container to delete" )
        else:

            # delete the item
            try:
                deleted_item = self._solarwinds.invoke  ( 
                                        "Orion.Container",
                                        "DeleteContainer",
                                        container [ 'results' ][ 0 ][ 'ContainerID' ]
                                    )
                print ( "Group: {} deleted!".format ( name ) )
            except Exception:
                print ( "Unable to delete container: {}".format ( name ) )

    def deleteDefinition  ( self , id ):

        '''
            Method name      : deleteDefinition
        
            Method Purpose   : To delete a definition from a group
        
            Parameters       :
                - id         : The id of the group
        
            Returns          : None
        '''

        currentDef = self._solarwinds.query (
                    """
                    SELECT
                        DefinitionID
                    FROM
                        Orion.ContainerMemberDefinition
                    WHERE
                        ContainerID='{}'
                    """.format ( id )
                                            
                )

        newDef = self._solarwinds.invoke (
                    'Orion.Container',
                    'DeleteDefinition',
                    currentDef [ 'results' ][ 0 ][ 'DefinitionID' ]

                )

    def getGroupInfo      ( self , name ):

        '''
            Method name    : getGroupInfo
        
            Method Purpose : To get the container ID of the group
        
            Parameters     :
                - name     : The name of the group
        
            Returns        : The results of the search
        '''

        # find the entity from the database
        try:
            entity_uri = self._solarwinds.query (   """
                                                    SELECT
                                                        c.Name,
                                                        c.ContainerID,
                                                        c.Uri
                                                    FROM
                                                        Orion.Container c
                                                    WHERE
                                                        c.Name
                                                    LIKE
                                                        '{}%'
                                                    """.format ( name.lower ( ) )
                                                )

        except requests.exceptions.HTTPError:
            print ( "Unable to find ID" )
            return "None"
        else:
            if entity_uri [ 'results' ] == []:
                return "None"
            else:
                return entity_uri

    def getGroupList      ( self , topLevelGroup ):

        '''
            Method name         : findGroupList
        
            Method Purpose      : To get the names of all the groups
        
            Parameters          :
                - topLevelGroup : The group that contains what you are looking for
        
            Returns             : The list of groups
        '''

        group_query = self._solarwinds.query (  """
                                                SELECT
                                                    c.Name
                                                FROM
                                                    Orion.Container t
                                                INNER JOIN
                                                    Orion.ContainerMembers c
                                                ON
                                                    t.ContainerID=c.ContainerID
                                                WHERE
                                                    t.Name='{}'
                                                """.format ( topLevelGroup )
                                             )

        # create the list based on query info
        groupList = []
        if len ( group_query [ 'results' ] ) > 0:
            groupList = [ item [ 'Name' ] for item in group_query [ 'results' ] ]
        return groupList

    def getNodeContain    ( self , node_name , existingGroups ):

        '''
            Method name    : findNodeContain
        
            Method Purpose : To find if a node is contained by a group
        
            Parameters           :  
                - node_name      : The node to check if it is contained
                - existingGroups : A list of all the groups created already
        
            Returns              :
                - True/False     : Whether there is a group containing node or not
                - GroupName/None : The name of the group of none
                - GroupID/0      : The id if it exists or 0 if none
        '''
        
        node_group  = self._solarwinds.query (  """"
                                                SELECT
                                                    c.Container.DisplayName as Name,
                                                    c.ContainerID           as ID,
                                                    c.Name                  as Member
                                                FROM
                                                    Orion.ContainerMembers c
                                                WHERE
                                                    c.Name LIKE '{}%'
                                                """.format ( node_name )
                                             )

        result_list = node_group [ 'results' ]
        if len ( result_list ) > 0:
            for group_holder in result_list:
                if group_holder [ 'Name' ] in existingGroups:
                    return True, group_holder [ 'Name' ], group_holder [ 'ID' ]
       
        # return that the group did not exist
        return False, "None", 0

    def queryGroupInfo    ( self , name  , days ):

        # test function for querying data
        result = self._solarwinds.query (   """
                                            SELECT
                                                Year    ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as Year,
                                                Month   ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as Month,
                                                Day     ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as Day,
                                                Hour    ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as Hour,
                                                Minute  ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as Minute,
                                                Second  ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as Second,
                                                WeekDay ( Tolocal ( r.ResponseTimeHistory.DateTime ) ) as WeekDay,
                                                r.ResponseTimeHistory.Availability as Available
                                            FROM
                                                Orion.Nodes r
                                            WHERE 
                                                r.Caption='{}' AND
                                                DayDiff ( Tolocal ( r.ResponseTimeHistory.DateTime ) , GetDate ( ) ) < {}
                                            ORDER BY
                                                r.ResponseTimeHistory.DateTime
                                            """.format ( name , days )
                                        )
        return result

# class GroupInfo       ( )

# class GroupMemberInfo ( )

# class NodeInfo        ( )


# class SolarProperties ( ABC ):

#     @abstractmethod
#     def getProperties    ( self , name ):
#         ''' To be implemented in child classes '''
#         pass

#     @abstractmethod
#     def createProperties ( self , name ):
#         ''' To be implemented in child classes '''
#         pass

#     @abstractmethod
#     def deleteProperties ( self , name ):
#         ''' To be implemented in child classes '''
#         pass
    
#     @abstractmethod
#     def updateProperties ( self , name ):
#         ''' To be implented in child classes '''
#         pass

# class NodeProperties  ( SolarProperties ):

#     '''
#           Class Name    : NodeProperties
#           Class Purpose : To allow for updating node properties
#     '''

#     def __init__ ( self , solarwinds_instance ):

#         '''
#             Method name               : __init__
        
#             Method Purpose            : To initialize a node instance
        
#             Parameters                :
#                 - solarwinds_instance : The solarwinds login object
        
#             Returns                   : NONE
#         '''

#         self._solarwinds = solarwinds_instance
#         self._data       = NodeInfo.queryAllProperties ( )

#     def getProperties    ( self , name ):

#     def createProperties ( self , name ):

#     def deleteProperties ( self , name ):

#     def updateProperites ( self , name ):

# class GroupProperties ( SolarProperties ):

#     '''
#         Class Name    : GroupProperties
#     	Class Purpose : To allow for updating group properties
#     '''

#     def __init__ ( self , group_name ):

#         '''
#             Method name    : __init__
        
#             Method Purpose : To initialize the group properties
        
#             Parameters     :
#                 - solarwinds_instance : The solarwinds login object
        
#             Returns        : NONE
#         '''

#         self._solarwinds = solarwinds_instance
#         self._data       = GroupInfo.queryAllProperties ( )

#     def getProperties    ( self , name ):

#     def createProperties ( self , name ):

#     def deleteProperties ( self , name ):

#     def updateProperites ( self , name ):
    
